# # excel_seed.py
# from __future__ import annotations

# from enum import Enum
# from pathlib import Path
# from typing import Any, Dict, List, Optional, Tuple
# from collections import defaultdict

# import openpyxl
# from pydantic import BaseModel, Field


# # ============================================================
# # Local lightweight schemas (avoid circular imports with flows.py)
# # ============================================================
# class FlowNodeType(str, Enum):
#     QUESTION = "QUESTION"
#     MESSAGE = "MESSAGE"
#     ALERT = "ALERT"


# class SeverityLevel(str, Enum):
#     GREEN = "GREEN"
#     AMBER = "AMBER"
#     RED = "RED"


# class FlowOptionIn(BaseModel):
#     display_order: int
#     label: str
#     value: str
#     severity: SeverityLevel
#     news2_score: int = 0
#     seriousness_points: int = 0
#     next_node_key: Optional[str] = None


# class FlowNodeIn(BaseModel):
#     node_key: str
#     node_type: FlowNodeType
#     category: int = 1

#     title: Optional[str] = None
#     body_text: str
#     help_text: Optional[str] = None

#     parent_node_key: Optional[str] = None
#     default_next_node_key: Optional[str] = None
#     auto_next_node_key: Optional[str] = None

#     ui_ack_required: bool = False
#     alert_severity: Optional[SeverityLevel] = None
#     notify_admin: bool = False

#     options: List[FlowOptionIn] = Field(default_factory=list)


# # ============================================================
# # Helpers
# # ============================================================
# def _as_str(x: Any) -> str:
#     if x is None:
#         return ""
#     return str(x).strip()


# def _to_int(x: Any) -> int:
#     if x is None:
#         return 0
#     try:
#         return int(float(str(x).strip()))
#     except Exception:
#         return 0


# def _norm_next_raw(x: Any) -> str:
#     s = _as_str(x)
#     if not s:
#         return "END"
#     if s.lower() == "end":
#         return "END"
#     return s


# def _norm_sev_from_rag(x: Any) -> SeverityLevel:
#     s = _as_str(x).lower()
#     if not s:
#         return SeverityLevel.GREEN
#     if "red" in s:
#         return SeverityLevel.RED
#     if "amber" in s or "ambar" in s:
#         return SeverityLevel.AMBER
#     return SeverityLevel.GREEN


# def _sev_from_points(points: int) -> SeverityLevel:
#     if points >= 100:
#         return SeverityLevel.RED
#     if points >= 30:
#         return SeverityLevel.AMBER
#     return SeverityLevel.GREEN


# # ============================================================
# # Category-safe remap
# # - Category 1: C1Q1..C1Qn
# # - Category 2: C2Q1..C2Qn
# # - Next to other category => END
# # ============================================================
# def _remap_by_category(nodes: List[FlowNodeIn]) -> tuple[str, List[FlowNodeIn]]:
#     if not nodes:
#         return "C2Q1", nodes

#     cat1 = [n for n in nodes if int(n.category) == 1]
#     cat2 = [n for n in nodes if int(n.category) == 2]

#     map1: Dict[str, str] = {n.node_key: f"C1Q{i}" for i, n in enumerate(cat1, start=1)}
#     map2: Dict[str, str] = {n.node_key: f"C2Q{i}" for i, n in enumerate(cat2, start=1)}

#     def get_cat_of_old_key(k: str) -> Optional[int]:
#         if k in map1:
#             return 1
#         if k in map2:
#             return 2
#         return None

#     def map_key(old_key: str) -> str:
#         if old_key in map1:
#             return map1[old_key]
#         if old_key in map2:
#             return map2[old_key]
#         # shouldn't happen, but safe:
#         return old_key

#     def map_next(src_cat: int, raw_next: Optional[str]) -> str:
#         if not raw_next:
#             return "END"
#         if raw_next == "END":
#             return "END"

#         t = raw_next.strip()

#         # exact match in either category?
#         dst_cat = get_cat_of_old_key(t)
#         if not dst_cat:
#             return "END"

#         # enforce same-category rule
#         if int(dst_cat) != int(src_cat):
#             return "END"

#         return map_key(t)

#     remapped: List[FlowNodeIn] = []

#     for n in nodes:
#         src_cat = int(n.category)
#         new_key = map_key(n.node_key)

#         nn = FlowNodeIn(
#             node_key=new_key,
#             node_type=n.node_type,
#             category=src_cat,
#             title=n.title,
#             body_text=n.body_text or "dummy",
#             help_text=n.help_text,
#             parent_node_key=None,  # no hierarchy
#             default_next_node_key=None,
#             auto_next_node_key=None,
#             ui_ack_required=bool(n.ui_ack_required),
#             alert_severity=n.alert_severity,
#             notify_admin=bool(n.notify_admin),
#             options=[],
#         )

#         # MESSAGE / ALERT => auto_next (must be same category or END)
#         if nn.node_type in (FlowNodeType.MESSAGE, FlowNodeType.ALERT):
#             nn.auto_next_node_key = map_next(src_cat, n.auto_next_node_key or "END")
#             if nn.node_type == FlowNodeType.ALERT and not nn.alert_severity:
#                 nn.alert_severity = SeverityLevel.GREEN
#             remapped.append(nn)
#             continue

#         # QUESTION => options nexts
#         opts: List[FlowOptionIn] = []
#         for i, o in enumerate(n.options or [], start=1):
#             opts.append(
#                 FlowOptionIn(
#                     display_order=i,
#                     label=o.label or "dummy",
#                     value=f"{new_key}_opt{i}",
#                     severity=o.severity or SeverityLevel.GREEN,
#                     news2_score=_to_int(o.news2_score),
#                     seriousness_points=_to_int(o.seriousness_points),
#                     next_node_key=map_next(src_cat, o.next_node_key or "END"),
#                 )
#             )

#         # ensure >=2 options
#         if len(opts) < 2:
#             opts.append(
#                 FlowOptionIn(
#                     display_order=len(opts) + 1,
#                     label="dummy",
#                     value=f"{new_key}_opt{len(opts) + 1}",
#                     severity=SeverityLevel.GREEN,
#                     news2_score=0,
#                     seriousness_points=0,
#                     next_node_key="END",
#                 )
#             )
#         nn.options = opts
#         remapped.append(nn)

#     # Start node: prefer category 1 start if exists else category 2
#     start = "C1Q1" if cat1 else ("C2Q1" if cat2 else remapped[0].node_key)
#     return start, remapped


# # ============================================================
# # Main: Excel -> (start_node_key, nodes)
# # ============================================================
# def build_nodes_from_excel(excel_path: Path) -> tuple[str, list[FlowNodeIn]]:
#     wb = openpyxl.load_workbook(excel_path, data_only=True)

#     ws_cat2 = wb["Colorectal symptoms and signs"]
#     ws_cat1 = wb["Clincial Obs- Colorectal"]

#     # ---------------- Category 2 ----------------
#     cat2_rows = list(ws_cat2.iter_rows(values_only=True))

#     key_to_question: Dict[str, str] = {}
#     key_to_option_rows: Dict[str, list[tuple[Any, Any, Any]]] = defaultdict(list)

#     cur_key = None
#     cur_q = None

#     for r in cat2_rows[2:]:
#         key = _as_str(r[0])
#         q = r[1]
#         resp = r[2]
#         nxt = r[4]
#         rag = r[5]

#         if key:
#             cur_key = key
#             cur_q = _as_str(q) or "dummy"

#         if not cur_key:
#             continue

#         if cur_key not in key_to_question:
#             key_to_question[cur_key] = cur_q or "dummy"

#         if resp is None and nxt is None and rag is None:
#             continue

#         key_to_option_rows[cur_key].append((resp, nxt, rag))

#     def _first_next_for_key(k: str) -> str:
#         for _resp, nxt, _rag in key_to_option_rows.get(k, []):
#             if _as_str(nxt):
#                 return _norm_next_raw(nxt)
#         return "END"

#     cat2_nodes: list[FlowNodeIn] = []

#     for k, qtext in key_to_question.items():
#         qtext = qtext or "dummy"
#         kl = k.lower()

#         if kl.startswith("instruction"):
#             cat2_nodes.append(
#                 FlowNodeIn(
#                     node_key=k,
#                     node_type=FlowNodeType.MESSAGE,
#                     category=2,
#                     body_text=qtext,
#                     ui_ack_required=True,
#                     auto_next_node_key=_first_next_for_key(k),
#                     options=[],
#                 )
#             )
#             continue

#         if kl.startswith("alert"):
#             cat2_nodes.append(
#                 FlowNodeIn(
#                     node_key=k,
#                     node_type=FlowNodeType.ALERT,
#                     category=2,
#                     body_text=qtext,
#                     ui_ack_required=True,
#                     alert_severity=SeverityLevel.GREEN,
#                     auto_next_node_key=_first_next_for_key(k),
#                     options=[],
#                 )
#             )
#             continue

#         opt_rows = key_to_option_rows.get(k, [])
#         options: list[FlowOptionIn] = []

#         for i, (resp, nxt, rag) in enumerate(opt_rows, start=1):
#             options.append(
#                 FlowOptionIn(
#                     display_order=i,
#                     label=_as_str(resp) or "dummy",
#                     value=f"{k}_opt{i}",  # temp
#                     severity=_norm_sev_from_rag(rag),
#                     news2_score=0,
#                     seriousness_points=0,
#                     next_node_key=_norm_next_raw(nxt),
#                 )
#             )

#         if len(options) < 2:
#             options.append(
#                 FlowOptionIn(
#                     display_order=len(options) + 1,
#                     label="dummy",
#                     value=f"{k}_opt{len(options) + 1}",
#                     severity=SeverityLevel.GREEN,
#                     news2_score=0,
#                     seriousness_points=0,
#                     next_node_key="END",
#                 )
#             )

#         cat2_nodes.append(
#             FlowNodeIn(
#                 node_key=k,
#                 node_type=FlowNodeType.QUESTION,
#                 category=2,
#                 body_text=qtext,
#                 options=options,
#             )
#         )

#     # ---------------- Category 1 ----------------
#     cat1_rows = list(ws_cat1.iter_rows(values_only=True))

#     num_to_shown: Dict[str, str] = {}
#     num_to_question: Dict[str, str] = {}
#     num_to_ans_rows: Dict[str, list[tuple[Any, Any, Any, Any]]] = defaultdict(list)

#     cur_num = None
#     cur_question = None
#     cur_shown = None

#     for r in cat1_rows[3:]:
#         num = _as_str(r[0])
#         shown = _as_str(r[1])
#         q = r[2]
#         ans = r[3]
#         news2 = r[4]
#         pts = r[5]
#         nxt = r[6]

#         if num:
#             cur_num = num
#             cur_shown = shown or ""
#             cur_question = _as_str(q) or "dummy"

#         if not cur_num:
#             continue

#         if cur_num not in num_to_question:
#             num_to_question[cur_num] = cur_question or "dummy"
#             num_to_shown[cur_num] = cur_shown or ""

#         if ans is None and nxt is None and news2 is None and pts is None:
#             continue

#         num_to_ans_rows[cur_num].append((ans, news2, pts, nxt))

#     def _first_next_cat1(n: str) -> str:
#         for ans, news2, pts, nxt in num_to_ans_rows.get(n, []):
#             if _as_str(nxt):
#                 return _norm_next_raw(nxt)
#         return "END"

#     cat1_nodes: list[FlowNodeIn] = []

#     for n, qtext in num_to_question.items():
#         shown = (num_to_shown.get(n, "") or "").strip().lower()
#         qtext = qtext or "dummy"

#         if shown == "hidden":
#             cat1_nodes.append(
#                 FlowNodeIn(
#                     node_key=n,
#                     node_type=FlowNodeType.MESSAGE,
#                     category=1,
#                     body_text=qtext,
#                     ui_ack_required=True,
#                     auto_next_node_key=_first_next_cat1(n),
#                     options=[],
#                 )
#             )
#             continue

#         options: list[FlowOptionIn] = []
#         rows = num_to_ans_rows.get(n, [])

#         for i, (ans, news2, pts, nxt) in enumerate(rows, start=1):
#             p = _to_int(pts)
#             options.append(
#                 FlowOptionIn(
#                     display_order=i,
#                     label=_as_str(ans) or "dummy",
#                     value=f"{n}_opt{i}",  # temp
#                     severity=_sev_from_points(p),
#                     news2_score=_to_int(news2),
#                     seriousness_points=p,
#                     next_node_key=_norm_next_raw(nxt),
#                 )
#             )

#         if len(options) < 2:
#             options.append(
#                 FlowOptionIn(
#                     display_order=len(options) + 1,
#                     label="dummy",
#                     value=f"{n}_opt{len(options) + 1}",
#                     severity=SeverityLevel.GREEN,
#                     news2_score=0,
#                     seriousness_points=0,
#                     next_node_key="END",
#                 )
#             )

#         cat1_nodes.append(
#             FlowNodeIn(
#                 node_key=n,
#                 node_type=FlowNodeType.QUESTION,
#                 category=1,
#                 body_text=qtext,
#                 options=options,
#             )
#         )

#     combined = cat1_nodes + cat2_nodes

#     if not combined:
#         combined = [
#             FlowNodeIn(
#                 node_key="dummy1",
#                 node_type=FlowNodeType.QUESTION,
#                 category=2,
#                 body_text="dummy",
#                 options=[
#                     FlowOptionIn(display_order=1, label="dummy", value="d1", severity=SeverityLevel.GREEN, next_node_key="END"),
#                     FlowOptionIn(display_order=2, label="dummy", value="d2", severity=SeverityLevel.GREEN, next_node_key="END"),
#                 ],
#             )
#         ]

#     # ✅ remap safely and enforce same-category next (else END)
#     start, remapped = _remap_by_category(combined)
#     return start, remapped

# excel_seed.py
from __future__ import annotations

from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict

import openpyxl
from pydantic import BaseModel, Field


# ============================================================
# Local lightweight schemas (avoid circular imports with flows.py)
# ============================================================
class FlowNodeType(str, Enum):
    QUESTION = "QUESTION"
    MESSAGE = "MESSAGE"
    ALERT = "ALERT"


class SeverityLevel(str, Enum):
    GREEN = "GREEN"
    AMBER = "AMBER"
    RED = "RED"


class FlowOptionIn(BaseModel):
    display_order: int
    label: str
    value: str
    severity: SeverityLevel
    news2_score: int = 0
    seriousness_points: int = 0
    next_node_key: Optional[str] = None


class FlowNodeIn(BaseModel):
    node_key: str
    node_type: FlowNodeType
    category: int = 1

    title: Optional[str] = None
    body_text: str
    help_text: Optional[str] = None

    parent_node_key: Optional[str] = None
    default_next_node_key: Optional[str] = None
    auto_next_node_key: Optional[str] = None

    ui_ack_required: bool = False
    alert_severity: Optional[SeverityLevel] = None
    notify_admin: bool = False

    options: List[FlowOptionIn] = Field(default_factory=list)


# ============================================================
# Helpers
# ============================================================
def _as_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _to_int(x: Any) -> int:
    if x is None:
        return 0
    try:
        return int(float(str(x).strip()))
    except Exception:
        return 0


def _norm_next_raw(x: Any) -> str:
    s = _as_str(x)
    if not s:
        return "END"
    if s.lower() == "end":
        return "END"
    return s


def _norm_sev_from_rag(x: Any) -> SeverityLevel:
    # Rule: blank RAG => GREEN
    s = _as_str(x).lower()
    if not s:
        return SeverityLevel.GREEN
    if "red" in s:
        return SeverityLevel.RED
    if "amber" in s or "ambar" in s:
        return SeverityLevel.AMBER
    return SeverityLevel.GREEN


def _sev_from_points(points: int) -> SeverityLevel:
    if points >= 100:
        return SeverityLevel.RED
    if points >= 30:
        return SeverityLevel.AMBER
    return SeverityLevel.GREEN


def _is_red_fill(cell) -> bool:
    """
    Detect if Excel cell has red-ish fill.
    Works best for explicit RGB fills (common in demos).
    """
    try:
        fill = getattr(cell, "fill", None)
        if not fill or not getattr(fill, "patternType", None):
            return False

        color = getattr(fill, "fgColor", None)
        if not color:
            return False

        rgb = None
        if getattr(color, "type", None) == "rgb":
            rgb = color.rgb  # e.g. "FFFF0000"
        else:
            # theme/indexed are harder; skip to avoid false positives
            return False

        if not rgb:
            return False

        rgb = str(rgb).upper()

        # Strong reds (common)
        if rgb.endswith("FF0000"):
            return True
        if rgb.endswith("CC0000"):
            return True
        if rgb.endswith("D00000"):
            return True
        if rgb.endswith("E00000"):
            return True

        return False
    except Exception:
        return False


# ============================================================
# Category-safe remap
# - Category 1: C1Q1..C1Qn
# - Category 2: C2Q1..C2Qn
# - Next to other category => END
# ============================================================
def _remap_by_category(nodes: List[FlowNodeIn]) -> tuple[str, List[FlowNodeIn]]:
    if not nodes:
        return "C2Q1", nodes

    cat1 = [n for n in nodes if int(n.category) == 1]
    cat2 = [n for n in nodes if int(n.category) == 2]

    map1: Dict[str, str] = {n.node_key: f"C1Q{i}" for i, n in enumerate(cat1, start=1)}
    map2: Dict[str, str] = {n.node_key: f"C2Q{i}" for i, n in enumerate(cat2, start=1)}

    def get_cat_of_old_key(k: str) -> Optional[int]:
        if k in map1:
            return 1
        if k in map2:
            return 2
        return None

    def map_key(old_key: str) -> str:
        if old_key in map1:
            return map1[old_key]
        if old_key in map2:
            return map2[old_key]
        return old_key

    def map_next(src_cat: int, raw_next: Optional[str]) -> str:
        if not raw_next:
            return "END"
        if raw_next == "END":
            return "END"

        t = raw_next.strip()

        dst_cat = get_cat_of_old_key(t)
        if not dst_cat:
            return "END"

        # enforce same-category rule
        if int(dst_cat) != int(src_cat):
            return "END"

        return map_key(t)

    remapped: List[FlowNodeIn] = []

    for n in nodes:
        src_cat = int(n.category)
        new_key = map_key(n.node_key)

        nn = FlowNodeIn(
            node_key=new_key,
            node_type=n.node_type,
            category=src_cat,
            title=n.title,
            body_text=n.body_text or "dummy",
            help_text=n.help_text,
            parent_node_key=None,  # no hierarchy
            default_next_node_key=None,
            auto_next_node_key=None,
            ui_ack_required=bool(n.ui_ack_required),
            alert_severity=n.alert_severity,
            notify_admin=bool(n.notify_admin),
            options=[],
        )

        if nn.node_type in (FlowNodeType.MESSAGE, FlowNodeType.ALERT):
            nn.auto_next_node_key = map_next(src_cat, n.auto_next_node_key or "END")
            if nn.node_type == FlowNodeType.ALERT and not nn.alert_severity:
                nn.alert_severity = SeverityLevel.GREEN
            remapped.append(nn)
            continue

        opts: List[FlowOptionIn] = []
        for i, o in enumerate(n.options or [], start=1):
            opts.append(
                FlowOptionIn(
                    display_order=i,
                    label=o.label or "dummy",
                    value=f"{new_key}_opt{i}",
                    severity=o.severity or SeverityLevel.GREEN,
                    news2_score=_to_int(o.news2_score),
                    seriousness_points=_to_int(o.seriousness_points),
                    next_node_key=map_next(src_cat, o.next_node_key or "END"),
                )
            )

        if len(opts) < 2:
            opts.append(
                FlowOptionIn(
                    display_order=len(opts) + 1,
                    label="dummy",
                    value=f"{new_key}_opt{len(opts) + 1}",
                    severity=SeverityLevel.GREEN,
                    news2_score=0,
                    seriousness_points=0,
                    next_node_key="END",
                )
            )

        nn.options = opts
        remapped.append(nn)

    start = "C1Q1" if cat1 else ("C2Q1" if cat2 else remapped[0].node_key)
    return start, remapped


# ============================================================
# Main: Excel -> (start_node_key, nodes)
# ============================================================
def build_nodes_from_excel(excel_path: Path) -> tuple[str, list[FlowNodeIn]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    ws_cat2 = wb["Colorectal symptoms and signs"]
    ws_cat1 = wb["Clincial Obs- Colorectal"]

    # ============================================================
    # CATEGORY 2 parsing (WITH RED ALERT FILL DETECTION)
    # A node_key, B Question, C Response, E Next, F RAG
    # ============================================================
    rows = list(ws_cat2.iter_rows(min_row=1, values_only=False))

    key_to_question: Dict[str, str] = {}
    key_to_alert_is_red: Dict[str, bool] = {}
    key_to_option_rows: Dict[str, list[tuple[Any, Any, Any]]] = defaultdict(list)

    cur_key = None
    cur_q = None

    # skip first two rows (title + headers)
    for row in rows[2:]:
        cell_key = row[0]  # col A
        cell_q = row[1]    # col B
        cell_resp = row[2] # col C
        cell_next = row[4] # col E
        cell_rag = row[5]  # col F

        key = _as_str(cell_key.value)
        q_val = cell_q.value
        resp = cell_resp.value
        nxt = cell_next.value
        rag = cell_rag.value

        if key:
            cur_key = key
            cur_q = _as_str(q_val) or "dummy"

            if cur_key.lower().startswith("alert"):
                key_to_alert_is_red[cur_key] = _is_red_fill(cell_key) or _is_red_fill(cell_q)

        if not cur_key:
            continue

        if cur_key not in key_to_question:
            key_to_question[cur_key] = cur_q or "dummy"

        if resp is None and nxt is None and rag is None:
            continue

        key_to_option_rows[cur_key].append((resp, nxt, rag))

    def _first_next_for_key(k: str) -> str:
        for _resp, nxt, _rag in key_to_option_rows.get(k, []):
            if _as_str(nxt):
                return _norm_next_raw(nxt)
        return "END"

    cat2_nodes: list[FlowNodeIn] = []

    for k, qtext in key_to_question.items():
        qtext = qtext or "dummy"
        kl = k.lower()

        if kl.startswith("instruction"):
            cat2_nodes.append(
                FlowNodeIn(
                    node_key=k,
                    node_type=FlowNodeType.MESSAGE,
                    category=2,
                    body_text=qtext,
                    ui_ack_required=True,
                    auto_next_node_key=_first_next_for_key(k),
                    options=[],
                )
            )
            continue

        if kl.startswith("alert"):
            is_red = bool(key_to_alert_is_red.get(k, False))
            cat2_nodes.append(
                FlowNodeIn(
                    node_key=k,
                    node_type=FlowNodeType.ALERT,
                    category=2,
                    body_text=qtext,
                    ui_ack_required=True,
                    alert_severity=SeverityLevel.RED if is_red else SeverityLevel.GREEN,
                    auto_next_node_key=_first_next_for_key(k),
                    options=[],
                )
            )
            continue

        opt_rows = key_to_option_rows.get(k, [])
        options: list[FlowOptionIn] = []

        for i, (resp, nxt, rag) in enumerate(opt_rows, start=1):
            options.append(
                FlowOptionIn(
                    display_order=i,
                    label=_as_str(resp) or "dummy",
                    value=f"{k}_opt{i}",  # temp
                    severity=_norm_sev_from_rag(rag),
                    news2_score=0,
                    seriousness_points=0,
                    next_node_key=_norm_next_raw(nxt),
                )
            )

        if len(options) < 2:
            options.append(
                FlowOptionIn(
                    display_order=len(options) + 1,
                    label="dummy",
                    value=f"{k}_opt{len(options) + 1}",
                    severity=SeverityLevel.GREEN,
                    news2_score=0,
                    seriousness_points=0,
                    next_node_key="END",
                )
            )

        cat2_nodes.append(
            FlowNodeIn(
                node_key=k,
                node_type=FlowNodeType.QUESTION,
                category=2,
                body_text=qtext,
                options=options,
            )
        )

    # ============================================================
    # CATEGORY 1 parsing
    # A Number, B Shown/Hidden, C Question, D Answer, E NEWS2, F Points, G Next
    # ============================================================
    cat1_rows = list(ws_cat1.iter_rows(values_only=True))

    num_to_shown: Dict[str, str] = {}
    num_to_question: Dict[str, str] = {}
    num_to_ans_rows: Dict[str, list[tuple[Any, Any, Any, Any]]] = defaultdict(list)

    cur_num = None
    cur_question = None
    cur_shown = None

    for r in cat1_rows[3:]:
        num = _as_str(r[0])
        shown = _as_str(r[1])
        q = r[2]
        ans = r[3]
        news2 = r[4]
        pts = r[5]
        nxt = r[6]

        if num:
            cur_num = num
            cur_shown = shown or ""
            cur_question = _as_str(q) or "dummy"

        if not cur_num:
            continue

        if cur_num not in num_to_question:
            num_to_question[cur_num] = cur_question or "dummy"
            num_to_shown[cur_num] = cur_shown or ""

        if ans is None and nxt is None and news2 is None and pts is None:
            continue

        num_to_ans_rows[cur_num].append((ans, news2, pts, nxt))

    def _first_next_cat1(n: str) -> str:
        for ans, news2, pts, nxt in num_to_ans_rows.get(n, []):
            if _as_str(nxt):
                return _norm_next_raw(nxt)
        return "END"

    cat1_nodes: list[FlowNodeIn] = []

    for n, qtext in num_to_question.items():
        shown = (num_to_shown.get(n, "") or "").strip().lower()
        qtext = qtext or "dummy"

        if shown == "hidden":
            cat1_nodes.append(
                FlowNodeIn(
                    node_key=n,
                    node_type=FlowNodeType.MESSAGE,
                    category=1,
                    body_text=qtext,
                    ui_ack_required=True,
                    auto_next_node_key=_first_next_cat1(n),
                    options=[],
                )
            )
            continue

        options: list[FlowOptionIn] = []
        rows = num_to_ans_rows.get(n, [])

        for i, (ans, news2, pts, nxt) in enumerate(rows, start=1):
            p = _to_int(pts)
            options.append(
                FlowOptionIn(
                    display_order=i,
                    label=_as_str(ans) or "dummy",
                    value=f"{n}_opt{i}",  # temp
                    severity=_sev_from_points(p),
                    news2_score=_to_int(news2),
                    seriousness_points=p,
                    next_node_key=_norm_next_raw(nxt),
                )
            )

        if len(options) < 2:
            options.append(
                FlowOptionIn(
                    display_order=len(options) + 1,
                    label="dummy",
                    value=f"{n}_opt{len(options) + 1}",
                    severity=SeverityLevel.GREEN,
                    news2_score=0,
                    seriousness_points=0,
                    next_node_key="END",
                )
            )

        cat1_nodes.append(
            FlowNodeIn(
                node_key=n,
                node_type=FlowNodeType.QUESTION,
                category=1,
                body_text=qtext,
                options=options,
            )
        )

    combined = cat1_nodes + cat2_nodes

    if not combined:
        combined = [
            FlowNodeIn(
                node_key="dummy1",
                node_type=FlowNodeType.QUESTION,
                category=2,
                body_text="dummy",
                options=[
                    FlowOptionIn(
                        display_order=1,
                        label="dummy",
                        value="d1",
                        severity=SeverityLevel.GREEN,
                        next_node_key="END",
                    ),
                    FlowOptionIn(
                        display_order=2,
                        label="dummy",
                        value="d2",
                        severity=SeverityLevel.GREEN,
                        next_node_key="END",
                    ),
                ],
            )
        ]

    # ✅ remap safely and enforce same-category next (else END)
    start, remapped = _remap_by_category(combined)
    return start, remapped
